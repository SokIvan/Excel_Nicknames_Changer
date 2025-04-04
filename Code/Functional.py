from libs import pygame,os,random,re,openpyxl as xl
from PIL import Image
from openpyxl.styles import Font, Alignment, PatternFill
#flg1 - менять ники
#flg2 - менять лидеров
#flg3 - менять часто бьют
#flg4 - вспомнить ники(старые)
#flg5 - вспомнить лидеры(старые)
#flg6 - вспомнить часто бьют(старые)
#flg7 - 1-менять, 0 - добавить
async def Parse(parse_mod,file):
    if parse_mod == "Default":
        table = file.split('\n')
        while table and table[-1] == '':
            table.pop()
        return table
    elif parse_mod == "VADIM_BOT":

        table = re.split(r'\nУчастники:\n|\nРуководство:\n|\nПомощники:\n|\nНаблюдатели:\n', file)

        lead_t = [item for item in table[1].split('\n') if item != '']
        user_t = [item for item in table[2].split('\n') if item != '']

        if len(lead_t) > 0:
            for i in range(len(lead_t)):
                lead_t[i] = lead_t[i].split(' ')[1]


        for i in range(len(user_t)):
            user_t[i] = user_t[i].split(' ')[1]

        table = [lead_t , user_t]
        return table
    else:
        return None
async def Process(folders, process_types):
    return [await Parse(t, open(f, encoding='utf-8').read()) for f, t in zip(folders, process_types)]

async def GENERATE_1(folders,process_types,flg_mas):

    Preprocess = await Process(folders,process_types)
    print("Preprocess",Preprocess)
    last_nicknames = Preprocess[3]
    last_leaders = Preprocess[4]   #Прошлый вход
    last_fighters = Preprocess[5]


    mas_nicknames = []
    mas_leaders = []    #вход
    mas_fighters = []

    mas_new_members = []
    mas_new_leaders = []    #обработка
    mas_new_fighters = []

    # with open('../Files/Last_nicknames', 'r') as f:
    #     read = f.read()
    #     last_nicknames = read.split('\n')
    #
    # with open('../Files/Last_leaders', 'r') as f:
    #     last_leaders = f.read().split('\n')
    #
    # with open('../Files/Last_most_valuable', 'r') as f:
    #     last_fighters = f.read().split('\n')

    if process_types[0] == "VADIM_BOT":
        mas_nicknames += Preprocess[0][0]
        mas_nicknames += Preprocess[0][1]
    elif process_types[0] == "Default":
        mas_nicknames += Preprocess[0]

    if process_types[1] == "VADIM_BOT":
        mas_leaders += Preprocess[1][0]
    elif process_types[1] == "Default":
        mas_leaders += Preprocess[1]

    if process_types[2] == "Default":
        mas_fighters += Preprocess[2]

    for i in mas_nicknames:
        if i not in last_nicknames:
            mas_new_members.append(i)

    for i in mas_leaders:
        if i not in last_leaders:
            mas_new_leaders.append(i)

    for i in mas_fighters:
        if i not in last_fighters:
            mas_new_fighters.append(i)


    if flg_mas[6]:
        ret_n = last_nicknames
        ret_l = last_leaders
        ret_f = last_fighters
        if flg_mas[0]:
            ret_n = mas_nicknames
        if flg_mas[1]:
            ret_l = mas_leaders
        if flg_mas[2]:
            ret_f = mas_fighters
        return [ret_n, ret_l, ret_f, [], [], []]
    else:
        ret_n = last_nicknames
        ret_l = last_leaders
        ret_f = last_fighters
        if flg_mas[0]:
            ret_n += mas_new_members
        if flg_mas[1]:
            ret_l += mas_new_leaders
        if flg_mas[2]:
            ret_f += mas_new_fighters

        return [ret_n,ret_l,ret_f,mas_new_members,mas_new_leaders,mas_new_fighters]

async def EXCEL_1(process_mas_, set_name, list_name):
    process_mas = [x.copy() for x in process_mas_]

    wb = xl.Workbook()
    ws = wb.active
    ws.title = list_name

    with open(f'../Files/Kriterii_Sets', 'r', encoding='utf-8') as f:
        c = f.read()
        set_krit_list = c.split('\n')[0].split('::::')[1].split(',,,,')
        krit_list = c.split('\n')[1].split('::::')[1].split(',,,,')

    index = set_krit_list.index(set_name) if set_name in set_krit_list else -1
    if index == -1:
        raise Exception('Set not found')

    krit_list = eval(krit_list[index])

    with open('../Files/kriteries', 'r', encoding='utf-8') as f:
        c = f.read()
        krit_names = c.split('\n')[0].split('::::')[1].split(',,,,')
        krit_amount = c.split('\n')[2].split('::::')[1].split(',,,,')
        for i in range(len(krit_amount)):
            krit_amount[i] = int(krit_amount[i])
        krit_symbols = c.split('\n')[3].split('::::')[1].split(',,,,')
        for i in range(len(krit_symbols)):
            krit_symbols[i] = eval(krit_symbols[i])
        krit_not = c.split('\n')[4].split('::::')[1].split(',,,,')
        for i in range(len(krit_not)):
            krit_not[i] = eval(krit_not[i])
        krit_and_or = c.split('\n')[5].split('::::')[1].split(',,,,')
        for i in range(len(krit_and_or)):
            krit_and_or[i] = eval(krit_and_or[i])

    kriterii_values_list = [[] for _ in range(len(krit_list))]

    for i in range(len(krit_names)):
        listik = []
        if krit_names[i] in krit_list:
            listik.append(krit_amount[i])
            listik.append(krit_symbols[i])
            listik.append(krit_not[i])
            listik.append(krit_and_or[i])
            index = krit_list.index(krit_names[i])
            kriterii_values_list[index] = listik

    with open('../Files/excel', 'r', encoding='utf-8') as f:
        excel_settings = f.read().split('\n')
        font_name = excel_settings[0].split(':')[-1]
        font_size = excel_settings[1].split(':')[-1]
        column_width = int(excel_settings[2].split(':')[-1])
        for i in range(1, 11):
            ws.column_dimensions[chr(i + 96)].width = column_width

    for i in range(len(krit_list)):
        cell = ws.cell(row=1, column=i + 1, value=krit_list[i])
        cell.font = Font(name=font_name, size=int(font_size), bold=True, color="000000", italic=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    mas_fill_cells = []
    check_mas = process_mas[0]
    for j in range(len(krit_list)):
        fill = []
        N = len(process_mas[0])
        k = 0
        for i in range(N):
            litstr = f"{check_mas[k]} "
            if await check_by_kriterii(check_mas[k], kriterii_values_list[j][0], kriterii_values_list[j][1], kriterii_values_list[j][2], kriterii_values_list[j][3]):
                litstr += "✅"
                fill.append(check_mas[k])
                check_mas.remove(check_mas[k])
                k -= 1
                N -= 1
            else:
                litstr += "❌"
            k += 1

        fill.sort(key=lambda x: [x.lower(), x])
        mas_fill_cells.append(fill)

    for i in range(len(mas_fill_cells)):
        for j in range(len(mas_fill_cells[i])):
            cell = ws.cell(row=j + 2, column=i + 1, value=mas_fill_cells[i][j])
            cell.font = Font(name=font_name, size=int(font_size), bold=True)
            if mas_fill_cells[i][j] in process_mas_[3]:
                cell.fill = PatternFill(start_color="c4fafa", fill_type="solid")
            if mas_fill_cells[i][j] in process_mas_[1]:
                cell.fill = PatternFill(start_color="c4f5a9", fill_type="solid")
                if mas_fill_cells[i][j] in process_mas_[4]:
                    cell.fill = PatternFill(start_color="a3ee76", fill_type="solid")
            elif mas_fill_cells[i][j] in process_mas_[2]:
                cell.fill = PatternFill(start_color="ffe0b5", fill_type="solid")
                if mas_fill_cells[i][j] in process_mas_[5]:
                    cell.fill = PatternFill(start_color="ffb54d", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(f'../Excel/{list_name}.xlsx')





def play_music(music_folder):
    # Инициализация pygame
    pygame.mixer.init()

    # Получаем список всех аудиофайлов в папке
    music_files = [f for f in os.listdir(music_folder) if f.endswith(('.mp3', '.wav'))]

    while True:
        # Выбираем случайный файл
        music_file = random.choice(music_files)
        music_path = os.path.join(music_folder, music_file)

        # Загружаем и воспроизводим музыку
        pygame.mixer.music.load(music_path)
        pygame.mixer.music.play()

        # Ждем, пока музыка не закончится
        while pygame.mixer.music.get_busy():
            pygame.time.Clock().tick(10)
# Функция для конвертации изображения в формат, читаемый PySimpleGUI
async def convert_image_for_sg(filename):
    try:
        with Image.open(filename) as img:
            img = img.resize((800, 400))
            img = img.convert('RGB')
            img.save(filename, format='PNG')
    except OSError:
        pass
async def choose_and_check_kriterii(kriterii_name):
    pass




#РАБОТАЕТ НЕПРАВИЛЬНО
async def check_by_kriterii(check_string, kriterii_amount=1, kriterii_symbols_mas=[], kriterii_not_mas=[],
                      kriterii_and_or_mas=[]):
    # Проверка на количество критериев
    if not (1 <= kriterii_amount <= 5):
        raise ValueError("kriterii_amount должно быть от 1 до 5")

    # Список для хранения результатов проверок
    results = []

    for i in range(kriterii_amount):
        symbol = kriterii_symbols_mas[i]
        not_flag = kriterii_not_mas[i]

        # Проверка наличия символа в строке
        if not_flag == 0:
            results.append(symbol in check_string)  # Если не отрицать, добавляем результат
        else:
            results.append(symbol not in check_string)  # Если отрицать, добавляем противоположный результат

    # Обработка связок
    final_result = results[0]
    for i in range(1, kriterii_amount):
        if kriterii_and_or_mas[i - 1] == 'and':
            final_result = final_result and results[i]
        elif kriterii_and_or_mas[i - 1] == 'or':
            final_result = final_result or results[i]

    return final_result















